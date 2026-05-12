import argparse
import csv
import datetime as dt
from pathlib import Path
from typing import Any

import openpyxl


def parse_date(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    s = str(v).strip()
    if not s:
        return None
    # "2024.04.16." -> "2024-04-16"
    if "." in s:
        s = s.replace(".", "-").strip("-")
    try:
        return dt.date.fromisoformat(s[:10]).isoformat()
    except ValueError:
        return None


def parse_int(v: Any) -> int | None:
    if v is None or str(v).strip() == "":
        return None
    try:
        return int(float(v))
    except ValueError:
        return None


def parse_float(v: Any) -> float | None:
    if v is None or str(v).strip() == "":
        return None
    try:
        return float(v)
    except ValueError:
        return None


def non_empty_row(values: list[Any]) -> bool:
    return any(v is not None and str(v).strip() != "" for v in values)


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def main() -> None:
    ap = argparse.ArgumentParser(description="Convert dashboard Excel to analytics CSV files.")
    ap.add_argument("--xlsx", required=True, help="Path to source .xlsx")
    ap.add_argument("--hospital-id", required=True, help="analytics hospital_id")
    ap.add_argument("--account-id", required=True, help="account_id for blog/place daily tables")
    ap.add_argument("--customer-id", required=True, help="customer_id for searchad daily table")
    ap.add_argument("--out-dir", default="scripts/analytics/out", help="Output directory")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    sheets = wb.worksheets
    if len(sheets) < 4:
        raise RuntimeError("Expected at least 4 worksheets: 블로그/플레이스/파워링크광고/플레이스광고")

    blog_ws, place_ws, power_ws, placead_ws = sheets[:4]

    blog_rows: list[dict[str, Any]] = []
    for row in blog_ws.iter_rows(min_row=2, values_only=True):
        if not non_empty_row(list(row[:4])):
            continue
        metric_date = parse_date(row[0])
        if not metric_date:
            continue
        blog_rows.append(
            {
                "account_id": args.account_id,
                "hospital_id": args.hospital_id,
                "hospital_name": "",
                "metric_date": metric_date,
                "blog_views": parse_int(row[1]) or 0,
                "blog_unique_visitors": parse_int(row[2]) or 0,
                "metadata": "{}",
            }
        )

    place_rows: list[dict[str, Any]] = []
    for row in place_ws.iter_rows(min_row=2, values_only=True):
        if not non_empty_row(list(row[:2])):
            continue
        metric_date = parse_date(row[0])
        if not metric_date:
            continue
        place_rows.append(
            {
                "account_id": args.account_id,
                "hospital_id": args.hospital_id,
                "hospital_name": "",
                "metric_date": metric_date,
                "smartplace_inflow": parse_int(row[1]) or 0,
                "metadata": "{}",
            }
        )

    power_rows: list[dict[str, Any]] = []
    # described columns: B 캠페인, C 광고그룹, D 날짜, E 노출수, F 클릭수, G 클릭율, H 평균CPC, I 총비용
    for row in power_ws.iter_rows(min_row=2, values_only=True):
        vals = list(row[:9])
        if not non_empty_row(vals):
            continue
        metric_date = parse_date(vals[3])
        if not metric_date:
            continue
        campaign_name = str(vals[1]).strip() if vals[1] is not None else ""
        adgroup_name = str(vals[2]).strip() if vals[2] is not None else ""
        if not campaign_name and not adgroup_name:
            continue
        power_rows.append(
            {
                "metric_date": metric_date,
                "hospital_id": args.hospital_id,
                "customer_id": args.customer_id,
                "campaign_id": "",
                "campaign_name": campaign_name,
                "adgroup_id": "",
                "adgroup_name": adgroup_name,
                "keyword_id": "",
                "impressions": parse_int(vals[4]) or 0,
                "clicks": parse_int(vals[5]) or 0,
                "cost": parse_float(vals[8]) or 0,
                "conversions": "",
                "conversion_value": "",
                "raw_payload": "{}",
            }
        )

    # placead sheet is same column layout as powerlink in this sample description.
    placead_rows: list[dict[str, Any]] = []
    for row in placead_ws.iter_rows(min_row=2, values_only=True):
        vals = list(row[:9])
        if not non_empty_row(vals):
            continue
        metric_date = parse_date(vals[3])
        if not metric_date:
            continue
        campaign_name = str(vals[1]).strip() if vals[1] is not None else ""
        adgroup_name = str(vals[2]).strip() if vals[2] is not None else ""
        if not campaign_name and not adgroup_name:
            continue
        placead_rows.append(
            {
                "metric_date": metric_date,
                "hospital_id": args.hospital_id,
                "customer_id": args.customer_id,
                "campaign_id": "",
                "campaign_name": campaign_name,
                "adgroup_id": "",
                "adgroup_name": adgroup_name,
                "keyword_id": "",
                "impressions": parse_int(vals[4]) or 0,
                "clicks": parse_int(vals[5]) or 0,
                "cost": parse_float(vals[8]) or 0,
                "conversions": "",
                "conversion_value": "",
                "raw_payload": "{\"ad_channel\":\"place_ad\"}",
            }
        )

    out_dir = Path(args.out_dir)
    write_csv(
        out_dir / "analytics_blog_daily_metrics.csv",
        ["account_id", "hospital_id", "hospital_name", "metric_date", "blog_views", "blog_unique_visitors", "metadata"],
        blog_rows,
    )
    write_csv(
        out_dir / "analytics_smartplace_daily_metrics.csv",
        ["account_id", "hospital_id", "hospital_name", "metric_date", "smartplace_inflow", "metadata"],
        place_rows,
    )
    write_csv(
        out_dir / "analytics_searchad_daily_metrics_powerlink.csv",
        [
            "metric_date",
            "hospital_id",
            "customer_id",
            "campaign_id",
            "campaign_name",
            "adgroup_id",
            "adgroup_name",
            "keyword_id",
            "impressions",
            "clicks",
            "cost",
            "conversions",
            "conversion_value",
            "raw_payload",
        ],
        power_rows,
    )
    write_csv(
        out_dir / "analytics_searchad_daily_metrics_placead.csv",
        [
            "metric_date",
            "hospital_id",
            "customer_id",
            "campaign_id",
            "campaign_name",
            "adgroup_id",
            "adgroup_name",
            "keyword_id",
            "impressions",
            "clicks",
            "cost",
            "conversions",
            "conversion_value",
            "raw_payload",
        ],
        placead_rows,
    )

    print(f"written: {out_dir}")
    print(f"- blog rows: {len(blog_rows)}")
    print(f"- place rows: {len(place_rows)}")
    print(f"- powerlink rows: {len(power_rows)}")
    print(f"- placead rows: {len(placead_rows)}")


if __name__ == "__main__":
    main()

